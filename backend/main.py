"""
Taxa Turism Manager — Backend FastAPI
Toate rutele API pentru gestionarea declarațiilor de taxă turistică.
"""

import os
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import List, Optional

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent / ".env")

import secrets
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from sqlmodel import Session, select

# Adaugă directorul curent în sys.path pentru importuri relative
sys.path.insert(0, str(Path(__file__).parent))

from database import engine, get_session, init_db
from models import Declaratie, Proprietate, Rezervare, StatusDeclaratie, Tranzactie
from services.booking_parser import parseaza_xls, importa_xls_complet
from services.folder_builder import construieste_folder
from services.pdf_generator import genereaza_pdf

BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = BASE_DIR / "output"

app = FastAPI(title="Taxa Turism Manager", version="1.0.0")

# ── HTTP Basic Auth ──────────────────────────────────────
_security = HTTPBasic(auto_error=False)
_APP_USER = os.environ.get("APP_USERNAME", "admin")
_APP_PASS = os.environ.get("APP_PASSWORD", "changeme")

def require_auth(credentials: HTTPBasicCredentials = Depends(_security)):
    """Verifică credențialele. Rutele publice (/api/fisa, /fisa) sunt excluse în middleware."""
    ok = credentials is not None and (
        secrets.compare_digest(credentials.username.encode(), _APP_USER.encode()) and
        secrets.compare_digest(credentials.password.encode(), _APP_PASS.encode())
    )
    if not ok:
        raise HTTPException(
            status_code=401,
            detail="Autentificare necesară",
            headers={"WWW-Authenticate": "Basic"},
        )

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def basic_auth_middleware(request: Request, call_next):
    """Protejează toate rutele cu Basic Auth, exceptând fișa publică a oaspeților."""
    public_prefixes = ("/api/fisa/", "/fisa/", "/assets/", "/favicon")
    path = request.url.path
    if any(path.startswith(p) for p in public_prefixes):
        return await call_next(request)

    # Verifică Authorization header
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Basic "):
        import base64
        try:
            decoded = base64.b64decode(auth[6:]).decode("utf-8")
            username, _, password = decoded.partition(":")
            user_ok = secrets.compare_digest(username.encode(), _APP_USER.encode())
            pass_ok = secrets.compare_digest(password.encode(), _APP_PASS.encode())
            if user_ok and pass_ok:
                return await call_next(request)
        except Exception:
            pass

    from fastapi.responses import Response
    return Response(
        content="Autentificare necesară",
        status_code=401,
        headers={"WWW-Authenticate": 'Basic realm="PropertyManager"'},
    )


@app.on_event("startup")
def on_startup():
    init_db()


# ──────────────────────────────────────────────
# Proprietăți
# ──────────────────────────────────────────────

@app.get("/api/proprietati")
def lista_proprietati(session: Session = Depends(get_session)):
    props = session.exec(select(Proprietate).where(Proprietate.activa == True)).all()
    return [
        {
            "id": p.id,
            "slug": p.slug,
            "nume": p.nume,
            "adresa": p.adresa,
            "sector": p.sector,
            "oras": p.oras,
            "autoritate": p.autoritate,
            "taxa_per_noapte": p.taxa_per_noapte,
        }
        for p in props
    ]


# ──────────────────────────────────────────────
# Declarații
# ──────────────────────────────────────────────

@app.get("/api/declaratii")
def lista_declaratii(
    an: Optional[int] = None,
    status: Optional[str] = None,
    proprietate_id: Optional[int] = None,
    session: Session = Depends(get_session),
):
    query = select(Declaratie)
    if an:
        query = query.where(Declaratie.an == an)
    if status:
        query = query.where(Declaratie.status == status)
    if proprietate_id:
        query = query.where(Declaratie.proprietate_id == proprietate_id)
    query = query.order_by(Declaratie.an.desc(), Declaratie.luna.desc())

    declaratii = session.exec(query).all()
    result = []
    for d in declaratii:
        prop = session.get(Proprietate, d.proprietate_id)
        result.append({
            "id": d.id,
            "proprietate_id": d.proprietate_id,
            "proprietate_nume": prop.nume if prop else None,
            "proprietate_slug": prop.slug if prop else None,
            "luna": d.luna,
            "an": d.an,
            "data_generare": d.data_generare.isoformat(),
            "total_nopti": d.total_nopti,
            "total_persoane_zile": d.total_persoane_zile,
            "taxa_totala": d.taxa_totala,
            "status": d.status,
            "folder_output": d.folder_output,
            "pdf_path": d.pdf_path,
        })
    return result


@app.get("/api/declaratii/{id}")
def detaliu_declaratie(id: int, session: Session = Depends(get_session)):
    decl = session.get(Declaratie, id)
    if not decl:
        raise HTTPException(status_code=404, detail="Declarație negăsită")

    prop = session.get(Proprietate, decl.proprietate_id)
    rezervari = session.exec(
        select(Rezervare).where(Rezervare.declaratie_id == id)
    ).all()

    return {
        "id": decl.id,
        "proprietate_id": decl.proprietate_id,
        "proprietate_nume": prop.nume if prop else None,
        "luna": decl.luna,
        "an": decl.an,
        "data_generare": decl.data_generare.isoformat(),
        "total_nopti": decl.total_nopti,
        "total_persoane_zile": decl.total_persoane_zile,
        "taxa_totala": decl.taxa_totala,
        "status": decl.status,
        "folder_output": decl.folder_output,
        "pdf_path": decl.pdf_path,
        "rezervari": [
            {
                "id": r.id,
                "booking_id": r.booking_id,
                "check_in": r.check_in.isoformat(),
                "check_out": r.check_out.isoformat(),
                "persoane": r.persoane,
                "nopti_in_luna": r.nopti_in_luna,
                "taxa_aferenta": r.taxa_aferenta,
            }
            for r in rezervari
        ],
    }


@app.post("/api/declaratii/preview")
async def preview_declaratie(
    xls_file: UploadFile = File(...),
    luna: int = Form(...),
    an: int = Form(...),
    proprietate_id: int = Form(...),
    session: Session = Depends(get_session),
):
    """
    Parsează XLS și returnează preview cu rezervări noi și duplicate.
    Nu salvează nimic în DB.
    """
    prop = session.get(Proprietate, proprietate_id)
    if not prop:
        raise HTTPException(status_code=404, detail="Proprietate negăsită")

    if not (1 <= luna <= 12):
        raise HTTPException(status_code=422, detail="Luna trebuie să fie între 1 și 12")

    xls_bytes = await xls_file.read()

    try:
        rezultat = parseaza_xls(
            xls_bytes=xls_bytes,
            luna=luna,
            an=an,
            taxa_per_noapte=prop.taxa_per_noapte,
            session=session,
            proprietate_id=proprietate_id,
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    if not rezultat["rezervari_noi"] and not rezultat["rezervari_duplicate"]:
        raise HTTPException(
            status_code=422,
            detail=f"Nu există rezervări valide în luna {luna}/{an} în fișierul uploadat.",
        )

    return {
        "proprietate_id": proprietate_id,
        "luna": luna,
        "an": an,
        **rezultat,
    }


@app.post("/api/declaratii/genereaza")
async def genereaza_declaratie(
    xls_file: UploadFile = File(...),
    luna: int = Form(...),
    an: int = Form(...),
    proprietate_id: int = Form(...),
    session: Session = Depends(get_session),
):
    """
    Parsează XLS, salvează rezervările noi în DB, generează PDF și construiește folderul.
    """
    prop = session.get(Proprietate, proprietate_id)
    if not prop:
        raise HTTPException(status_code=404, detail="Proprietate negăsită")

    if not (1 <= luna <= 12):
        raise HTTPException(status_code=422, detail="Luna trebuie să fie între 1 și 12")

    # Verifică dacă există deja o declarație pentru această lună
    existenta = session.exec(
        select(Declaratie).where(
            Declaratie.proprietate_id == proprietate_id,
            Declaratie.luna == luna,
            Declaratie.an == an,
        )
    ).first()
    if existenta:
        raise HTTPException(
            status_code=409,
            detail=f"Există deja o declarație pentru {luna}/{an} (ID: {existenta.id}, status: {existenta.status}).",
        )

    xls_bytes = await xls_file.read()

    try:
        rezultat = parseaza_xls(
            xls_bytes=xls_bytes,
            luna=luna,
            an=an,
            taxa_per_noapte=prop.taxa_per_noapte,
            session=session,
            proprietate_id=proprietate_id,
        )
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    if not rezultat["rezervari_noi"]:
        raise HTTPException(
            status_code=422,
            detail="Toate rezervările din XLS sunt deja declarate. Nu se generează declarație nouă.",
        )

    # Salvează declarația
    decl = Declaratie(
        proprietate_id=proprietate_id,
        luna=luna,
        an=an,
        total_nopti=rezultat["total_nopti"],
        total_persoane_zile=rezultat["total_persoane_zile"],
        taxa_totala=rezultat["taxa_totala"],
        status=StatusDeclaratie.generat,
    )
    session.add(decl)
    session.flush()  # obține decl.id fără commit final

    # Salvează rezervările noi
    from datetime import date
    for r in rezultat["rezervari_noi"]:
        rez = Rezervare(
            declaratie_id=decl.id,
            proprietate_id=proprietate_id,
            booking_id=r["booking_id"],
            check_in=date.fromisoformat(r["check_in"]),
            check_out=date.fromisoformat(r["check_out"]),
            persoane=r["persoane"],
            nopti_in_luna=r["nopti_in_luna"],
            taxa_aferenta=r["taxa_aferenta"],
        )
        session.add(rez)

    # Actualizează declaratie_id pe rezervarile importate corespunzătoare
    from models import RezervaraImportata
    for r in rezultat["rezervari_noi"]:
        rim = session.exec(
            select(RezervaraImportata).where(RezervaraImportata.booking_id == r["booking_id"])
        ).first()
        if rim:
            rim.declaratie_id = decl.id
            session.add(rim)

    # Generează PDF
    if not Path(prop.template_pdf).exists():
        raise HTTPException(
            status_code=500,
            detail=f"Template PDF negăsit: {prop.template_pdf}",
        )

    pdf_filename = f"declaratie_taxa_{luna:02d}_{an}.pdf"
    pdf_tmp = Path(tempfile.mkdtemp()) / pdf_filename

    completate = genereaza_pdf(
        template_path=prop.template_pdf,
        output_path=str(pdf_tmp),
        luna=luna,
        an=an,
        total_nopti=rezultat["total_nopti"],
        total_persoane_zile=rezultat["total_persoane_zile"],
        taxa_totala=rezultat["taxa_totala"],
    )

    # Construiește folderul output
    documente_statice_dir = str(
        BASE_DIR / "data" / "proprietati" / prop.slug / "documente_statice"
    )

    folder_path = construieste_folder(
        proprietate_slug=prop.slug,
        luna=luna,
        an=an,
        pdf_declaratie_path=str(pdf_tmp),
        documente_statice_dir=documente_statice_dir,
        output_base_dir=str(OUTPUT_DIR),
    )

    pdf_final = str(Path(folder_path) / pdf_filename)

    # Actualizează declarația cu căile generate
    decl.pdf_path = pdf_final
    decl.folder_output = folder_path
    decl.status = StatusDeclaratie.generat
    session.add(decl)
    session.commit()
    session.refresh(decl)

    return {
        "id": decl.id,
        "proprietate_id": proprietate_id,
        "luna": luna,
        "an": an,
        "total_nopti": rezultat["total_nopti"],
        "total_persoane_zile": rezultat["total_persoane_zile"],
        "taxa_totala": rezultat["taxa_totala"],
        "status": decl.status,
        "pdf_path": pdf_final,
        "folder_output": folder_path,
        "rezervari_salvate": len(rezultat["rezervari_noi"]),
        "rezervari_duplicate_excluse": len(rezultat["rezervari_duplicate"]),
        "campuri_pdf_completate": completate,
    }


@app.post("/api/declaratii/genereaza-din-import")
def genereaza_din_import(
    proprietate_id: int = Form(...),
    luna: int = Form(...),
    an: int = Form(...),
    session: Session = Depends(get_session),
):
    """
    Generează declarație direct din rezervaraimportata — fără re-upload XLS.
    Folosește rezervările din luna/an care nu au încă o declarație.
    """
    from datetime import date
    from models import RezervaraImportata
    import calendar as cal_mod

    prop = session.get(Proprietate, proprietate_id)
    if not prop:
        raise HTTPException(status_code=404, detail="Proprietate negăsită")

    if not (1 <= luna <= 12):
        raise HTTPException(status_code=422, detail="Luna trebuie să fie între 1 și 12")

    # Verifică conflict lună
    existenta = session.exec(
        select(Declaratie).where(
            Declaratie.proprietate_id == proprietate_id,
            Declaratie.luna == luna,
            Declaratie.an == an,
        )
    ).first()
    if existenta:
        raise HTTPException(
            status_code=409,
            detail=f"Există deja o declarație pentru {luna}/{an} (ID: {existenta.id}, status: {existenta.status.value}).",
        )

    # Găsește rezervările cu check-out în luna respectivă care nu sunt declarate
    toate = session.exec(
        select(RezervaraImportata).where(
            RezervaraImportata.proprietate_id == proprietate_id,
            RezervaraImportata.declaratie_id == None,
        )
    ).all()

    rezervari_lunii = []
    for r in toate:
        if r.check_out.month == luna and r.check_out.year == an:
            n = (r.check_out - r.check_in).days
            rezervari_lunii.append((r, n))

    if not rezervari_lunii:
        raise HTTPException(
            status_code=422,
            detail=f"Nu există rezervări importate nedeclarate pentru {luna}/{an}. Importă mai întâi un XLS.",
        )

    total_nopti = sum(n for _, n in rezervari_lunii)
    total_pz    = sum(n * r.persoane for r, n in rezervari_lunii)
    taxa_totala = total_pz * prop.taxa_per_noapte

    # Salvează declarația
    decl = Declaratie(
        proprietate_id=proprietate_id,
        luna=luna, an=an,
        total_nopti=total_nopti,
        total_persoane_zile=total_pz,
        taxa_totala=taxa_totala,
        status=StatusDeclaratie.generat,
    )
    session.add(decl)
    session.flush()

    # Salvează rezervările în tabelul Rezervare + leagă RezervaraImportata
    for r, n in rezervari_lunii:
        session.add(Rezervare(
            declaratie_id=decl.id,
            proprietate_id=proprietate_id,
            booking_id=r.booking_id,
            check_in=r.check_in,
            check_out=r.check_out,
            persoane=r.persoane,
            nopti_in_luna=n,
            taxa_aferenta=n * r.persoane * prop.taxa_per_noapte,
        ))
        r.declaratie_id = decl.id
        session.add(r)

    # Generează PDF
    if not Path(prop.template_pdf).exists():
        raise HTTPException(status_code=500, detail=f"Template PDF negăsit: {prop.template_pdf}")

    pdf_filename = f"declaratie_taxa_{luna:02d}_{an}.pdf"
    pdf_tmp = Path(tempfile.mkdtemp()) / pdf_filename
    completate = genereaza_pdf(
        template_path=prop.template_pdf,
        output_path=str(pdf_tmp),
        luna=luna, an=an,
        total_nopti=total_nopti,
        total_persoane_zile=total_pz,
        taxa_totala=taxa_totala,
    )

    # Construiește folderul
    folder_path = construieste_folder(
        proprietate_slug=prop.slug,
        luna=luna, an=an,
        pdf_declaratie_path=str(pdf_tmp),
        documente_statice_dir=str(BASE_DIR / "data" / "proprietati" / prop.slug / "documente_statice"),
        output_base_dir=str(OUTPUT_DIR),
    )

    pdf_final = str(Path(folder_path) / pdf_filename)
    decl.pdf_path = pdf_final
    decl.folder_output = folder_path
    decl.status = StatusDeclaratie.generat
    session.add(decl)
    session.commit()
    session.refresh(decl)

    return {
        "id": decl.id,
        "proprietate_id": proprietate_id,
        "luna": luna, "an": an,
        "total_nopti": total_nopti,
        "total_persoane_zile": total_pz,
        "taxa_totala": taxa_totala,
        "status": decl.status,
        "pdf_path": pdf_final,
        "folder_output": folder_path,
        "rezervari_salvate": len(rezervari_lunii),
        "campuri_pdf_completate": completate,
    }


@app.post("/api/rezervari/scan")
async def scan_xls(
    xls_file: UploadFile = File(...),
    proprietate_id: int = Form(...),
    session: Session = Depends(get_session),
):
    """
    Importă toate rezervările din XLS în DB (persistent) și
    returnează lista grupată pe luni cu statusul declarației.
    """
    prop = session.get(Proprietate, proprietate_id)
    if not prop:
        raise HTTPException(status_code=404, detail="Proprietate negăsită")

    xls_bytes = await xls_file.read()
    try:
        stats = importa_xls_complet(xls_bytes, proprietate_id, session)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    _sync_fise(proprietate_id, session)

    rezultat = _grupuri_din_db(proprietate_id, session)
    rezultat["import_stats"] = stats
    return rezultat


@app.get("/api/rezervari")
def lista_rezervari(
    proprietate_id: int,
    session: Session = Depends(get_session),
):
    prop = session.get(Proprietate, proprietate_id)
    if not prop:
        raise HTTPException(status_code=404, detail="Proprietate negăsită")
    return _grupuri_din_db(proprietate_id, session)


@app.get("/api/curatenie")
def lista_curatenie(
    proprietate_id: int,
    luna: Optional[int] = None,
    an: Optional[int] = None,
    session: Session = Depends(get_session),
):
    from models import RezervaraImportata
    from datetime import date as date_cls, timedelta
    import calendar

    azi = date_cls.today()
    if luna is None:
        luna = azi.month
    if an is None:
        an = azi.year

    # Luăm rezervările relevante: cele care au checkout în luna selectată
    # + rezervările adiacente (pentru a calcula intervalele corect)
    # Strategie: luăm toate rezervările proprietății sortate după check_out
    toate = session.exec(
        select(RezervaraImportata)
        .where(RezervaraImportata.proprietate_id == proprietate_id)
        .order_by(RezervaraImportata.check_out, RezervaraImportata.check_in)
    ).all()

    if not toate:
        return {"intervale": [], "luna": luna, "an": an}

    # Deduplică și sortează după check_out
    vazute = set()
    unice = []
    for r in toate:
        if r.booking_id not in vazute:
            vazute.add(r.booking_id)
            unice.append(r)
    unice.sort(key=lambda r: (r.check_out, r.check_in))

    # Calculează intervalele de curățenie
    prima_zi_luna = date_cls(an, luna, 1)
    ultima_zi_luna = date_cls(an, luna, calendar.monthrange(an, luna)[1])

    intervale = []
    for i, rez in enumerate(unice):
        urmatoarea = unice[i + 1] if i + 1 < len(unice) else None

        checkout = rez.check_out
        checkin_urmator = urmatoarea.check_in if urmatoarea else None

        # Filtru: checkout-ul trebuie să fie în luna selectată
        if not (prima_zi_luna <= checkout <= ultima_zi_luna):
            continue

        if checkin_urmator:
            zile = (checkin_urmator - checkout).days
            urgent = zile == 0
        else:
            zile = None
            urgent = False

        intervale.append({
            "checkout": checkout.isoformat(),
            "guest_out": rez.nume_turist,
            "booking_out": rez.booking_id,
            "sursa_out": getattr(rez, "sursa", None) or "booking",
            "checkin_urmator": checkin_urmator.isoformat() if checkin_urmator else None,
            "guest_in": urmatoarea.nume_turist if urmatoarea else None,
            "booking_in": urmatoarea.booking_id if urmatoarea else None,
            "sursa_in": (getattr(urmatoarea, "sursa", None) or "booking") if urmatoarea else None,
            "zile_disponibile": zile,
            "urgent": urgent,
        })

    return {"intervale": intervale, "luna": luna, "an": an}


@app.delete("/api/rezervari")
def sterge_rezervari_luna(
    proprietate_id: int,
    luna: int,
    an: int,
    session: Session = Depends(get_session),
):
    """
    Șterge toate rezervările importate pentru o lună.
    Refuză dacă există o declarație generată pentru luna respectivă.
    """
    from models import RezervaraImportata
    import calendar as cal_mod
    from datetime import date, timedelta

    decl = session.exec(
        select(Declaratie).where(
            Declaratie.proprietate_id == proprietate_id,
            Declaratie.luna == luna,
            Declaratie.an == an,
        )
    ).first()
    if decl:
        raise HTTPException(
            status_code=409,
            detail=f"Există o declarație pentru {luna}/{an} (ID: {decl.id}, status: {decl.status.value}). Șterge mai întâi declarația.",
        )

    toate = session.exec(
        select(RezervaraImportata).where(
            RezervaraImportata.proprietate_id == proprietate_id
        )
    ).all()

    sterse = 0
    for r in toate:
        # Rezervarea aparține lunii dacă check-out e în luna respectivă
        if r.check_out.month == luna and r.check_out.year == an:
            fisa = session.exec(
                select(FisaOaspete).where(FisaOaspete.booking_id == r.booking_id)
            ).first()
            if fisa:
                session.delete(fisa)
            session.delete(r)
            sterse += 1

    session.commit()
    return {"sterse": sterse, "luna": luna, "an": an}


@app.post("/api/rezervari/manual")
def adauga_rezervare_manuala(
    proprietate_id: int = Form(...),
    check_in: str = Form(...),
    check_out: str = Form(...),
    persoane: int = Form(...),
    pret_platit: Optional[float] = Form(None),
    nume_turist: Optional[str] = Form(None),
    numar_rezervare: Optional[str] = Form(None),
    session: Session = Depends(get_session),
):
    """Adaugă o rezervare manuală (Airbnb). Validează suprapunerile de date."""
    from datetime import date as date_cls
    from models import RezervaraImportata
    import uuid as uuid_mod

    prop = session.get(Proprietate, proprietate_id)
    if not prop:
        raise HTTPException(status_code=404, detail="Proprietate negăsită")

    ci = date_cls.fromisoformat(check_in)
    co = date_cls.fromisoformat(check_out)
    if co <= ci:
        raise HTTPException(status_code=422, detail="Check-out trebuie să fie după check-in.")

    # Verifică suprapuneri
    toate = session.exec(
        select(RezervaraImportata).where(
            RezervaraImportata.proprietate_id == proprietate_id
        )
    ).all()
    for r in toate:
        if r.check_in < co and r.check_out > ci:
            raise HTTPException(
                status_code=409,
                detail=f"Suprapunere cu rezervarea {r.booking_id} ({r.check_in} – {r.check_out}).",
            )

    if numar_rezervare and numar_rezervare.strip():
        booking_id = f"AIRBNB-{numar_rezervare.strip()}"
    else:
        booking_id = f"AIRBNB-{uuid_mod.uuid4().hex[:8].upper()}"
    rez = RezervaraImportata(
        booking_id=booking_id,
        proprietate_id=proprietate_id,
        check_in=ci,
        check_out=co,
        persoane=persoane,
        pret_platit=pret_platit,
        nume_turist=nume_turist,
        sursa="airbnb",
    )
    session.add(rez)
    session.commit()
    return {"booking_id": booking_id, "message": "Rezervare adăugată."}


@app.delete("/api/rezervari/manual/{booking_id}")
def sterge_rezervare_manuala(
    booking_id: str,
    session: Session = Depends(get_session),
):
    """Șterge o rezervare manuală (airbnb/manual). Nu permite ștergerea celor din Booking.com."""
    from models import RezervaraImportata

    rez = session.exec(
        select(RezervaraImportata).where(RezervaraImportata.booking_id == booking_id)
    ).first()
    if not rez:
        raise HTTPException(status_code=404, detail="Rezervare negăsită.")
    sursa = getattr(rez, "sursa", None) or "booking"
    if sursa == "booking":
        raise HTTPException(status_code=403, detail="Nu poți șterge rezervări importate din Booking.com individual.")
    if rez.declaratie_id:
        raise HTTPException(status_code=409, detail="Rezervarea e deja declarată. Șterge mai întâi declarația.")
    session.delete(rez)
    session.commit()
    return {"deleted": booking_id}


@app.delete("/api/rezervari/{booking_id}")
def sterge_rezervare(
    booking_id: str,
    session: Session = Depends(get_session),
):
    """Șterge orice rezervare (Booking sau manual) care nu e încă declarată."""
    from models import RezervaraImportata

    rez = session.exec(
        select(RezervaraImportata).where(RezervaraImportata.booking_id == booking_id)
    ).first()
    if not rez:
        raise HTTPException(status_code=404, detail="Rezervare negăsită.")
    if rez.declaratie_id:
        raise HTTPException(status_code=409, detail="Rezervarea face parte dintr-o declarație deja generată. Șterge mai întâi declarația.")
    session.delete(rez)
    session.commit()
    return {"deleted": booking_id}


def _grupuri_din_db(proprietate_id: int, session: Session) -> dict:
    """Construiește răspunsul grupat pe luni din tabelul rezervaraimportata.
    O rezervare aparține lunii datei de check-out."""
    from models import RezervaraImportata

    prop = session.get(Proprietate, proprietate_id)
    taxa_per_noapte = prop.taxa_per_noapte if prop else 10.0

    rezervari = session.exec(
        select(RezervaraImportata)
        .where(RezervaraImportata.proprietate_id == proprietate_id)
        .order_by(RezervaraImportata.check_out)
    ).all()

    grupuri: dict[tuple, dict] = {}

    for r in rezervari:
        ci, co = r.check_in, r.check_out
        nopti = (co - ci).days  # total nopți sejur
        key = (co.month, co.year)
        if key not in grupuri:
            m, y = co.month, co.year
            decl = session.exec(
                select(Declaratie).where(
                    Declaratie.proprietate_id == proprietate_id,
                    Declaratie.luna == m,
                    Declaratie.an == y,
                )
            ).first()
            grupuri[key] = {
                "rezervari": [],
                "declaratie_id": decl.id if decl else None,
                "status_declaratie": decl.status.value if decl else None,
                "taxa_totala": decl.taxa_totala if decl else None,
            }
        grupuri[key]["rezervari"].append({
            "booking_id": r.booking_id,
            "check_in": ci.isoformat(),
            "check_out": co.isoformat(),
            "persoane": r.persoane,
            "nopti_in_luna": nopti,
            "nume_turist": r.nume_turist,
            "pret_platit": r.pret_platit,
            "sursa": getattr(r, "sursa", None) or "booking",
        })

    luni_list = sorted(
        [{"luna": l, "an": a, **g} for (l, a), g in grupuri.items()],
        key=lambda x: (x["an"], x["luna"])
    )

    # Compute taxa_totala for undeclared months using property rate
    for g in luni_list:
        if g["taxa_totala"] is None:
            total_pz = sum(r["nopti_in_luna"] * r["persoane"] for r in g["rezervari"])
            g["taxa_totala"] = round(total_pz * taxa_per_noapte, 2)

    return {"proprietate_id": proprietate_id, "luni": luni_list, "taxa_per_noapte": taxa_per_noapte}


def _date_range(ci, co):
    from datetime import timedelta
    d = ci
    while d < co:
        yield d
        d += timedelta(days=1)


@app.patch("/api/declaratii/{id}/status")
def actualizeaza_status(
    id: int,
    status: str,
    session: Session = Depends(get_session),
):
    decl = session.get(Declaratie, id)
    if not decl:
        raise HTTPException(status_code=404, detail="Declarație negăsită")
    try:
        decl.status = StatusDeclaratie(status)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Status invalid: {status}. Valid: generat, depus")
    session.add(decl)
    session.commit()
    return {"id": id, "status": decl.status}


@app.get("/api/declaratii/{id}/download")
def descarca_pdf(id: int, session: Session = Depends(get_session)):
    decl = session.get(Declaratie, id)
    if not decl:
        raise HTTPException(status_code=404, detail="Declarație negăsită")
    if not decl.pdf_path or not Path(decl.pdf_path).exists():
        raise HTTPException(status_code=404, detail="PDF-ul declarației nu a fost generat sau a fost șters.")
    return FileResponse(
        path=decl.pdf_path,
        media_type="application/pdf",
        filename=Path(decl.pdf_path).name,
    )


@app.get("/api/declaratii/{id}/folder")
def descarca_folder(id: int, session: Session = Depends(get_session)):
    import zipfile
    import io
    from fastapi.responses import StreamingResponse

    decl = session.get(Declaratie, id)
    if not decl:
        raise HTTPException(status_code=404, detail="Declarație negăsită")
    if not decl.folder_output or not Path(decl.folder_output).exists():
        raise HTTPException(status_code=404, detail="Folderul output nu există.")

    folder = Path(decl.folder_output)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in folder.iterdir():
            if f.is_file():
                zf.write(f, f.name)
    buf.seek(0)
    zip_name = f"declaratie_{id}.zip"
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={zip_name}"},
    )


@app.post("/api/declaratii/{id}/regenereaza")
def regenereaza_declaratie(id: int, session: Session = Depends(get_session)):
    """Regenerează PDF și folderul pentru o declarație existentă din datele din DB."""
    decl = session.get(Declaratie, id)
    if not decl:
        raise HTTPException(status_code=404, detail="Declarație negăsită")

    prop = session.get(Proprietate, decl.proprietate_id)
    if not prop:
        raise HTTPException(status_code=404, detail="Proprietate negăsită")

    if not Path(prop.template_pdf).exists():
        raise HTTPException(status_code=500, detail=f"Template PDF negăsit: {prop.template_pdf}")

    luna, an = decl.luna, decl.an
    pdf_filename = f"declaratie_taxa_{luna:02d}_{an}.pdf"
    pdf_tmp = Path(tempfile.mkdtemp()) / pdf_filename

    genereaza_pdf(
        template_path=prop.template_pdf,
        output_path=str(pdf_tmp),
        luna=luna, an=an,
        total_nopti=decl.total_nopti,
        total_persoane_zile=decl.total_persoane_zile,
        taxa_totala=decl.taxa_totala,
    )

    folder_path = construieste_folder(
        proprietate_slug=prop.slug,
        luna=luna, an=an,
        pdf_declaratie_path=str(pdf_tmp),
        documente_statice_dir=str(BASE_DIR / "data" / "proprietati" / prop.slug / "documente_statice"),
        output_base_dir=str(OUTPUT_DIR),
    )

    pdf_final = str(Path(folder_path) / pdf_filename)
    decl.pdf_path = pdf_final
    decl.folder_output = folder_path
    session.add(decl)
    session.commit()

    return {"ok": True, "folder_output": folder_path, "pdf_path": pdf_final}


@app.delete("/api/declaratii/{id}")
def sterge_declaratie(id: int, session: Session = Depends(get_session)):
    decl = session.get(Declaratie, id)
    if not decl:
        raise HTTPException(status_code=404, detail="Declarație negăsită")

    # Șterge toate fișierele din folderul output, lasă folderul gol
    if decl.folder_output:
        try:
            folder = Path(decl.folder_output)
            if folder.exists() and folder.is_dir():
                for f in folder.iterdir():
                    if f.is_file():
                        f.unlink()
        except Exception:
            pass

    # Curăță declaratie_id pe RezervaraImportata asociate
    from models import RezervaraImportata
    rezervari_importate = session.exec(
        select(RezervaraImportata).where(RezervaraImportata.declaratie_id == id)
    ).all()
    for ri in rezervari_importate:
        ri.declaratie_id = None
        session.add(ri)

    # Șterge rezervările asociate
    rezervari = session.exec(select(Rezervare).where(Rezervare.declaratie_id == id)).all()
    for r in rezervari:
        session.delete(r)

    session.delete(decl)
    session.commit()
    return {"deleted": True, "id": id}


# ──────────────────────────────────────────────
# Fișe oaspeți
# ──────────────────────────────────────────────

from models import FisaOaspete, StatusFisa

MESAJ_RO = """Bună ziua, {nume}!

Vă rugăm să completați fișa de înregistrare înainte de sosire, accesând link-ul de mai jos:

{link}

Detalii rezervare:
• Nr. rezervare: {booking_id}
• Check-in: {check_in}
• Check-out: {check_out}

Completarea durează aproximativ 2 minute și este obligatorie pentru check-in.

Cu drag,
Dan"""

MESAJ_EN = """Dear {nume},

Please complete the guest registration form before your arrival by clicking the link below:

{link}

Booking details:
• Booking number: {booking_id}
• Check-in: {check_in}
• Check-out: {check_out}

It only takes about 2 minutes and is mandatory for check-in.

Kind regards,
Dan"""

MESAJ_BILINGV = """Dear {nume},

Please complete the guest registration form before your arrival by clicking the link below:

{link}

Booking details:
• Booking number: {booking_id}
• Check-in: {check_in}
• Check-out: {check_out}

It only takes about 2 minutes and is mandatory for check-in.

Kind regards,
Dan

---

Bună ziua, {nume}!

Vă rugăm să completați fișa de înregistrare înainte de sosire, accesând link-ul de mai jos:

{link}

Detalii rezervare:
• Nr. rezervare: {booking_id}
• Check-in: {check_in}
• Check-out: {check_out}

Completarea durează aproximativ 2 minute și este obligatorie pentru check-in.

Cu drag,
Dan"""

SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY", "")
NOTIFICARE_DESTINATARI = ["danmugur.niculescu@gmail.com", "radu.niculescu2005@gmail.com"]
NOTIFICARE_SENDER = "info@dnrentals.eu"


def trimite_notificare_fisa(fisa) -> None:
    """Trimite email de notificare către proprietar când un turist completează fișa."""
    if not SENDGRID_API_KEY:
        print("⚠️  SENDGRID_API_KEY lipsă — emailul de notificare nu s-a trimis.")
        return
    try:
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail, To

        sex_label = {"M": "Masculin", "F": "Feminin"}.get(fisa.sex or "", fisa.sex or "—")
        ci = fisa.check_in.strftime("%d.%m.%Y") if fisa.check_in else "—"
        co = fisa.check_out.strftime("%d.%m.%Y") if fisa.check_out else "—"
        dn = fisa.data_nasterii.strftime("%d.%m.%Y") if fisa.data_nasterii else "—"

        def rand(label, val, bg="#ffffff"):
            return (
                f'<tr style="background:{bg}">'
                f'<td style="padding:8px 12px;font-weight:700;width:40%;color:#374151">{label}</td>'
                f'<td style="padding:8px 12px;color:#1a3a6b">{val}</td>'
                f"</tr>"
            )

        corp_html = f"""<!DOCTYPE html><html><body>
        <div style="font-family:Arial,sans-serif;max-width:580px;margin:0 auto">
          <div style="background:#1a3a6b;padding:20px 28px;border-radius:10px 10px 0 0">
            <h2 style="color:#fff;margin:0;font-size:18px">Fisa completata</h2>
            <p style="color:#a5b4fc;margin:4px 0 0;font-size:13px">Property Management · DNRentals</p>
          </div>
          <div style="background:#f8faff;padding:20px 28px;border:1px solid #e0e7ff;border-top:none">
            <p style="font-size:14px;color:#374151;margin-top:0">
              Turistul <strong>{fisa.nume_turist or "—"}</strong> a completat fisa de inregistrare.
            </p>
            <table style="width:100%;border-collapse:collapse;font-size:13px;border:1px solid #e0e7ff;border-radius:8px;overflow:hidden">
              {rand("Nr. rezervare", fisa.booking_id or "—", "#eef2ff")}
              {rand("Nume complet", fisa.nume_turist or "—")}
              {rand("Check-in", ci, "#eef2ff")}
              {rand("Check-out", co)}
              {rand("Sex", sex_label, "#eef2ff")}
              {rand("Data nasterii", dn)}
              {rand("Cetatenie", fisa.cetatenie or "—", "#eef2ff")}
              {rand("Document", f"{fisa.tip_document or '—'} {fisa.serie_numar or '—'} ({fisa.tara_emitenta or '—'})")}
              {rand("Domiciliu", fisa.domiciliu or "—", "#eef2ff")}
            </table>
            <p style="font-size:11px;color:#9ca3af;margin-bottom:0;margin-top:16px">
              info@dnrentals.eu · notificare automata
            </p>
          </div>
        </div>
        </body></html>"""

        subiect = f"Fisa completata: {fisa.nume_turist or fisa.booking_id} ({ci} - {co})"

        message = Mail(
            from_email=NOTIFICARE_SENDER,
            to_emails=[To(addr) for addr in NOTIFICARE_DESTINATARI],
            subject=subiect,
            html_content=corp_html,
        )
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        response = sg.send(message)
        print(f"✉️  Email notificare trimis — status {response.status_code}")
    except Exception as exc:
        print(f"⚠️  Eroare trimitere email notificare: {exc}")


def _sync_fise(proprietate_id: int, session: Session):
    """Creează/actualizează FisaOaspete pentru toate rezervările din proprietate."""
    from models import RezervaraImportata
    rezervari = session.exec(
        select(RezervaraImportata).where(RezervaraImportata.proprietate_id == proprietate_id)
    ).all()
    for r in rezervari:
        fisa = session.exec(
            select(FisaOaspete).where(FisaOaspete.booking_id == r.booking_id)
        ).first()
        if fisa:
            fisa.nume_turist = r.nume_turist
            fisa.email = r.email
            fisa.telefon = r.telefon
            fisa.check_in = r.check_in
            fisa.check_out = r.check_out
            session.add(fisa)
        else:
            session.add(FisaOaspete(
                booking_id=r.booking_id,
                proprietate_id=proprietate_id,
                check_in=r.check_in,
                check_out=r.check_out,
                nume_turist=r.nume_turist,
                email=r.email,
                telefon=r.telefon,
            ))
    session.commit()


@app.post("/api/fise/import")
async def importa_fise(
    xls_file: UploadFile = File(...),
    proprietate_id: int = Form(...),
    session: Session = Depends(get_session),
):
    """Importă rezervările din XLS și creează/actualizează fișele de oaspeți."""
    from services.booking_parser import importa_xls_complet
    from models import RezervaraImportata

    xls_bytes = await xls_file.read()
    importa_xls_complet(xls_bytes, proprietate_id, session)

    rezervari = session.exec(
        select(RezervaraImportata)
        .where(RezervaraImportata.proprietate_id == proprietate_id)
        .order_by(RezervaraImportata.check_in.desc())
    ).all()

    noi = actualizate = 0
    for r in rezervari:
        fisa = session.exec(
            select(FisaOaspete).where(FisaOaspete.booking_id == r.booking_id)
        ).first()
        if fisa:
            fisa.nume_turist = r.nume_turist
            fisa.email = r.email
            fisa.telefon = r.telefon
            fisa.check_in = r.check_in
            fisa.check_out = r.check_out
            session.add(fisa)
            actualizate += 1
        else:
            session.add(FisaOaspete(
                booking_id=r.booking_id,
                proprietate_id=proprietate_id,
                check_in=r.check_in,
                check_out=r.check_out,
                nume_turist=r.nume_turist,
                email=r.email,
                telefon=r.telefon,
            ))
            noi += 1
    session.commit()
    return {"noi": noi, "actualizate": actualizate}


@app.get("/api/fise")
def lista_fise(
    proprietate_id: int,
    session: Session = Depends(get_session),
):
    fise = session.exec(
        select(FisaOaspete)
        .where(FisaOaspete.proprietate_id == proprietate_id)
        .order_by(FisaOaspete.check_in.desc())
    ).all()
    return [_fisa_dict(f) for f in fise]


@app.patch("/api/fise/{id}/reset")
def reseteaza_fisa(id: int, session: Session = Depends(get_session)):
    """Șterge datele completate de turist, păstrează fișa."""
    fisa = session.get(FisaOaspete, id)
    if not fisa:
        raise HTTPException(status_code=404, detail="Fișă negăsită")
    fisa.prenume = None
    fisa.sex = None
    fisa.data_nasterii = None
    fisa.cetatenie = None
    fisa.tip_document = None
    fisa.serie_numar = None
    fisa.tara_emitenta = None
    fisa.domiciliu = None
    fisa.confirmare_date = False
    fisa.semnatura_nume = None
    fisa.semnatura_img = None
    fisa.completat_la = None
    fisa.status = StatusFisa.netrimis
    session.add(fisa)
    session.commit()
    return _fisa_dict(fisa)


@app.delete("/api/fise/{id}")
def sterge_fisa(id: int, session: Session = Depends(get_session)):
    """Șterge o fișă de oaspete."""
    fisa = session.get(FisaOaspete, id)
    if not fisa:
        raise HTTPException(status_code=404, detail="Fișă negăsită")
    session.delete(fisa)
    session.commit()
    return {"ok": True}


@app.patch("/api/fise/{id}/status")
def actualizeaza_status_fisa(
    id: int,
    status: str,
    session: Session = Depends(get_session),
):
    fisa = session.get(FisaOaspete, id)
    if not fisa:
        raise HTTPException(status_code=404, detail="Fișă negăsită")
    try:
        fisa.status = StatusFisa(status)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Status invalid: {status}")
    if status == "trimis" and not fisa.trimis_la:
        from datetime import datetime
        fisa.trimis_la = datetime.now()
    session.add(fisa)
    session.commit()
    return _fisa_dict(fisa)


@app.get("/api/fise/{id}/mesaj")
def genereaza_mesaj(
    id: int,
    request: Request,
    limba: str = "ro",
    session: Session = Depends(get_session),
):
    fisa = session.get(FisaOaspete, id)
    if not fisa:
        raise HTTPException(status_code=404, detail="Fișă negăsită")
    base = str(request.base_url).rstrip("/")
    link = f"{base}/fisa/{fisa.token}"
    if limba == "ro":
        template = MESAJ_RO
        nume_default = "Oaspete"
    elif limba == "en":
        template = MESAJ_EN
        nume_default = "Guest"
    else:  # bilingv
        template = MESAJ_BILINGV
        nume_default = "Guest"
    mesaj = template.format(
        nume=fisa.nume_turist or nume_default,
        link=link,
        booking_id=fisa.booking_id,
        check_in=fisa.check_in.strftime("%d.%m.%Y"),
        check_out=fisa.check_out.strftime("%d.%m.%Y"),
    )
    return {"mesaj": mesaj, "link": link}


@app.get("/api/fisa/{token}")
def get_fisa_publica(token: str, session: Session = Depends(get_session)):
    fisa = session.exec(select(FisaOaspete).where(FisaOaspete.token == token)).first()
    if not fisa:
        raise HTTPException(status_code=404, detail="Link invalid sau expirat.")
    return _fisa_dict(fisa)


@app.post("/api/fisa/{token}")
def submit_fisa_publica(token: str, body: dict, session: Session = Depends(get_session)):
    from datetime import datetime, date as date_cls
    fisa = session.exec(select(FisaOaspete).where(FisaOaspete.token == token)).first()
    if not fisa:
        raise HTTPException(status_code=404, detail="Link invalid sau expirat.")
    if fisa.status == StatusFisa.completat:
        raise HTTPException(status_code=409, detail="Fișa a fost deja completată.")

    fields = ["prenume", "sex", "cetatenie", "tip_document", "serie_numar",
              "tara_emitenta", "domiciliu", "semnatura_nume", "semnatura_img"]
    for f in fields:
        if f in body and body[f] is not None:
            setattr(fisa, f, body[f])

    if "data_nasterii" in body and body["data_nasterii"]:
        fisa.data_nasterii = date_cls.fromisoformat(body["data_nasterii"])

    fisa.confirmare_date = bool(body.get("confirmare_date", False))
    fisa.status = StatusFisa.completat
    fisa.completat_la = datetime.now()
    session.add(fisa)
    session.commit()
    session.refresh(fisa)
    trimite_notificare_fisa(fisa)
    return {"ok": True}


def _fisa_dict(f: FisaOaspete) -> dict:
    return {
        "id": f.id,
        "token": f.token,
        "booking_id": f.booking_id,
        "proprietate_id": f.proprietate_id,
        "check_in": f.check_in.isoformat(),
        "check_out": f.check_out.isoformat(),
        "nume_turist": f.nume_turist,
        "email": f.email,
        "telefon": f.telefon,
        "status": f.status,
        "creat_la": f.creat_la.isoformat(),
        "trimis_la": f.trimis_la.isoformat() if f.trimis_la else None,
        "completat_la": f.completat_la.isoformat() if f.completat_la else None,
        "prenume": f.prenume,
        "sex": f.sex,
        "data_nasterii": f.data_nasterii.isoformat() if f.data_nasterii else None,
        "cetatenie": f.cetatenie,
        "tip_document": f.tip_document,
        "serie_numar": f.serie_numar,
        "tara_emitenta": f.tara_emitenta,
        "domiciliu": f.domiciliu,
        "confirmare_date": f.confirmare_date,
        "semnatura_nume": f.semnatura_nume,
        "semnatura_img": f.semnatura_img,
    }


@app.get("/api/fise/export")
def export_fise_excel(
    proprietate_id: int,
    data_de: Optional[str] = None,
    data_pana: Optional[str] = None,
    session: Session = Depends(get_session),
):
    from datetime import date as date_cls
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    import io

    query = select(FisaOaspete).where(FisaOaspete.proprietate_id == proprietate_id)
    if data_de:
        query = query.where(FisaOaspete.check_in >= date_cls.fromisoformat(data_de))
    if data_pana:
        query = query.where(FisaOaspete.check_in <= date_cls.fromisoformat(data_pana))
    query = query.order_by(FisaOaspete.check_in)
    fise = session.exec(query).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fișe oaspeți"

    headers = [
        "Nr. rezervare", "Nume", "Prenume", "Sex", "Data nașterii", "Cetățenie",
        "Tip document", "Serie/Nr.", "Țara emitentă", "Domiciliu",
        "Check-in", "Check-out", "Email", "Telefon",
        "Status", "Completat la",
    ]

    header_fill = PatternFill("solid", fgColor="1A3A6B")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    for row, f in enumerate(fise, 2):
        vals = [
            f.booking_id,
            (f.nume_turist or "").split(" ", 1)[-1] if f.nume_turist and " " in (f.nume_turist or "") else f.nume_turist,
            f.prenume,
            "Bărbat" if f.sex == "M" else ("Femeie" if f.sex == "F" else f.sex),
            f.data_nasterii.strftime("%d.%m.%Y") if f.data_nasterii else "",
            f.cetatenie,
            f.tip_document,
            f.serie_numar,
            f.tara_emitenta,
            f.domiciliu,
            f.check_in.strftime("%d.%m.%Y") if f.check_in else "",
            f.check_out.strftime("%d.%m.%Y") if f.check_out else "",
            f.email,
            f.telefon,
            f.status.value if f.status else "",
            f.completat_la.strftime("%d.%m.%Y %H:%M") if f.completat_la else "",
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val or "")
        if row % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F0F4FF")

    col_widths = [16, 18, 14, 8, 14, 14, 12, 14, 14, 30, 11, 11, 24, 14, 11, 16]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"fise_oaspeti_{data_de or 'all'}_{data_pana or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ──────────────────────────────────────────────
# Financiar — extras de cont ING
# ──────────────────────────────────────────────

def _parse_extras_pdf(pdf_bytes: bytes) -> list[dict]:
    """
    Parsează un extras de cont ING România (PDF) și returnează tranzacțiile de tip Incasare.

    Format ING: fiecare tranzacție are:
      - Linie header: "DD luna YYYY  Incasare  <credit>  <balanta>"
      - Sub-linie: "Data: DD-MM-YYYY"
      - Sub-linie: "Ordonator:BOOKING.COM BV"
      - Sub-linie: "Referinta:<numar>"
    Tranzacțiile de tip Debit (Cumparare POS etc.) nu au Referinta și sunt ignorate.
    """
    import pdfplumber
    import re
    import io
    from datetime import date as date_cls

    # Extrage tot textul din PDF
    lines = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                lines.extend(t.split("\n"))

    tranzactii = []

    # Parcurge liniile; când găsim "Incasare" cu sume, colectăm blocul următor
    i = 0
    while i < len(lines):
        line = lines[i].strip()

        # Linie header Incasare: conține cuvântul "Incasare" și cel puțin două sume
        if "Incasare" in line:
            # Extrage toate sumele numerice din linie (format românesc: 1.234,56)
            sume = re.findall(r'\d{1,3}(?:\.\d{3})*,\d{2}', line)
            if len(sume) >= 1:
                # Prima sumă = credit (a doua ar fi balanta)
                credit_str = sume[0].replace(".", "").replace(",", ".")
                try:
                    credit = float(credit_str)
                except ValueError:
                    i += 1
                    continue

                # Caută în următoarele ~10 linii: Data:, Ordonator:, Referinta:
                data_val = None
                referinta_val = None
                ordonator_val = "BOOKING.COM BV"

                j = i + 1
                while j < len(lines) and j < i + 12:
                    sub = lines[j].strip()

                    m_data = re.match(r'Data:\s*(\d{2}-\d{2}-\d{4})', sub)
                    if m_data:
                        d = m_data.group(1).split("-")  # [DD, MM, YYYY]
                        try:
                            data_val = date_cls(int(d[2]), int(d[1]), int(d[0]))
                        except ValueError:
                            pass

                    m_ref = re.match(r'Referinta:(\d+)', sub)
                    if m_ref:
                        referinta_val = m_ref.group(1)

                    m_ord = re.match(r'Ordonator:(.+)', sub)
                    if m_ord:
                        ordonator_val = m_ord.group(1).strip()

                    j += 1

                if data_val and referinta_val and credit > 0:
                    tranzactii.append({
                        "referinta": referinta_val,
                        "data": data_val,
                        "suma": credit,
                        "ordonator": ordonator_val,
                    })

        i += 1

    return tranzactii


@app.post("/api/financiar/upload")
async def upload_extras(
    pdf_file: UploadFile = File(...),
    session: Session = Depends(get_session),
):
    """Parsează PDF extras de cont și salvează tranzacțiile (upsert după referinta)."""
    if not pdf_file.filename or not pdf_file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=422, detail="Fișierul trebuie să fie PDF.")

    pdf_bytes = await pdf_file.read()
    try:
        tranzactii_raw = _parse_extras_pdf(pdf_bytes)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Eroare la parsarea PDF: {e}")

    if not tranzactii_raw:
        raise HTTPException(
            status_code=422,
            detail="Nu s-au găsit tranzacții în PDF. Verifică formatul fișierului.",
        )

    noi = 0
    duplicate = 0
    for t in tranzactii_raw:
        existing = session.exec(
            select(Tranzactie).where(Tranzactie.referinta == t["referinta"])
        ).first()
        if existing:
            duplicate += 1
            continue
        session.add(Tranzactie(
            referinta=t["referinta"],
            data=t["data"],
            suma=t["suma"],
            ordonator=t.get("ordonator", "BOOKING.COM BV"),
        ))
        noi += 1

    session.commit()
    return {"noi": noi, "duplicate": duplicate, "total_gasite": len(tranzactii_raw)}


@app.get("/api/financiar")
def lista_tranzactii(
    an: Optional[int] = None,
    session: Session = Depends(get_session),
):
    """Returnează tranzacțiile grupate pe luni cu Taxa T din declarații. Filtrare după an."""
    from datetime import date as date_cls

    if an is None:
        an = date_cls.today().year

    query = select(Tranzactie).where(
        Tranzactie.data >= date_cls(an, 1, 1),
        Tranzactie.data <= date_cls(an, 12, 31),
    ).order_by(Tranzactie.data)

    tranzactii = session.exec(query).all()

    # Grupare pe luni
    grupuri: dict[tuple, list] = {}
    for t in tranzactii:
        key = (t.data.year, t.data.month)
        if key not in grupuri:
            grupuri[key] = []
        grupuri[key].append({
            "id": t.id,
            "referinta": t.referinta,
            "data": t.data.isoformat(),
            "suma": t.suma,
            "ordonator": t.ordonator,
        })

    # Calculează taxa T per (an, luna) din rezervarile importate
    # (nopti × persoane × taxa_per_noapte, grupate după check_out) — identic cu Rezervări.jsx
    from models import RezervaraImportata

    props = session.exec(select(Proprietate)).all()
    taxa_per_prop = {p.id: p.taxa_per_noapte for p in props}

    toate_rez = session.exec(select(RezervaraImportata)).all()
    taxa_per_luna: dict[tuple, float] = {}
    for r in toate_rez:
        co = r.check_out
        key = (co.year, co.month)
        nopti = (co - r.check_in).days
        taxa_noapte = taxa_per_prop.get(r.proprietate_id, 10.0)
        taxa_per_luna[key] = taxa_per_luna.get(key, 0.0) + nopti * r.persoane * taxa_noapte

    COTA_IMPOZIT = 0.07  # 7%

    luni_list = []
    total_incasari = 0.0
    total_taxa_t = 0.0
    for (y, m), items in sorted(grupuri.items(), reverse=True):
        subtotal = round(sum(i["suma"] for i in items), 2)
        taxa_t = round(taxa_per_luna.get((y, m), 0.0), 2)
        val_impozabila = round(subtotal - taxa_t, 2)
        impozit = round(val_impozabila * COTA_IMPOZIT, 2)
        total_incasari += subtotal
        total_taxa_t += taxa_t
        luni_list.append({
            "an": y,
            "luna": m,
            "tranzactii": items,
            "subtotal": subtotal,
            "taxa_t": taxa_t,
            "val_impozabila": val_impozabila,
            "impozit": impozit,
        })

    total_incasari = round(total_incasari, 2)
    total_taxa_t = round(total_taxa_t, 2)
    total_val_impozabila = round(total_incasari - total_taxa_t, 2)
    total_impozit = round(total_val_impozabila * COTA_IMPOZIT, 2)

    # Returnează și anii disponibili pentru filtru
    ani_query = session.exec(select(Tranzactie)).all()
    ani = sorted({t.data.year for t in ani_query}, reverse=True)

    return {
        "an": an,
        "luni": luni_list,
        "total_incasari": total_incasari,
        "total_taxa_t": total_taxa_t,
        "total_val_impozabila": total_val_impozabila,
        "total_impozit": total_impozit,
        "ani_disponibili": ani,
    }


@app.delete("/api/financiar/{id}")
def sterge_tranzactie(id: int, session: Session = Depends(get_session)):
    """Șterge o tranzacție individual."""
    t = session.get(Tranzactie, id)
    if not t:
        raise HTTPException(status_code=404, detail="Tranzacție negăsită")
    session.delete(t)
    session.commit()
    return {"ok": True}


# ──────────────────────────────────────────────
# Servire frontend React (producție)
# ──────────────────────────────────────────────

FRONTEND_DIST = Path(__file__).parent.parent / "frontend" / "dist"

if FRONTEND_DIST.exists():
    app.mount("/assets", StaticFiles(directory=str(FRONTEND_DIST / "assets")), name="assets")

    @app.get("/{full_path:path}", response_class=HTMLResponse, include_in_schema=False)
    async def serve_frontend(full_path: str):
        index = FRONTEND_DIST / "index.html"
        return HTMLResponse(index.read_text(encoding="utf-8"))
